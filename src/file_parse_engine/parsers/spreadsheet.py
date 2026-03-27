"""Spreadsheet parsers (XLSX / CSV / TSV) — direct conversion to Markdown tables.

XLSX enhancements:
- Formula semantics: annotate cells with formula comments
- Merged cell expansion: fill merged ranges with the merged value
- Smart table splitting: split large tables by category rows
"""

from __future__ import annotations

import csv
from pathlib import Path

from file_parse_engine.models import ParsedDocument, ParsedPage
from file_parse_engine.parsers import register
from file_parse_engine.parsers.base import BaseParser
from file_parse_engine.utils.logger import get_logger

logger = get_logger("parsers.spreadsheet")

# Max rows before triggering smart split
_SPLIT_THRESHOLD = 50


def _rows_to_markdown_table(rows: list[list[str]], *, title: str = "") -> str:
    """Convert a 2D list of strings to a Markdown table."""
    if not rows:
        return ""

    parts: list[str] = []
    if title:
        parts.append(f"## {title}")
        parts.append("")

    col_count = max(len(r) for r in rows)
    normalized = [r + [""] * (col_count - len(r)) for r in rows]

    header = normalized[0]
    parts.append("| " + " | ".join(header) + " |")
    parts.append("| " + " | ".join("---" for _ in header) + " |")

    for row in normalized[1:]:
        parts.append("| " + " | ".join(row) + " |")

    return "\n".join(parts)


# ------------------------------------------------------------------
# Formula helpers
# ------------------------------------------------------------------

_FORMULA_LABELS = {
    "SUM": "sum",
    "AVERAGE": "avg",
    "COUNT": "count",
    "MAX": "max",
    "MIN": "min",
    "VLOOKUP": "lookup",
    "HLOOKUP": "lookup",
    "IF": "conditional",
    "SUMIF": "conditional sum",
    "COUNTIF": "conditional count",
    "INDEX": "index",
    "MATCH": "match",
}


def _formula_annotation(formula: str) -> str:
    """Convert an Excel formula to a short semantic annotation."""
    if not formula or not formula.startswith("="):
        return ""

    f_upper = formula[1:].strip().upper()

    for func, label in _FORMULA_LABELS.items():
        if f_upper.startswith(func + "("):
            return f" <!-- {label}: {formula} -->"

    # Generic formula
    return f" <!-- formula: {formula} -->"


# ------------------------------------------------------------------
# Merged cell helpers
# ------------------------------------------------------------------

def _build_merge_map(ws) -> dict[tuple[int, int], str]:
    """Build a map of (row, col) → merged value for all merged ranges."""
    merge_map: dict[tuple[int, int], str] = {}

    for merge_range in ws.merged_cells.ranges:
        # Get the value from the top-left cell of the merge
        top_left = ws.cell(merge_range.min_row, merge_range.min_col)
        value = str(top_left.value) if top_left.value is not None else ""

        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merge_map[(row, col)] = value

    return merge_map


# ------------------------------------------------------------------
# Smart split helpers
# ------------------------------------------------------------------

def _is_category_row(cells: list[str], col_count: int) -> bool:
    """Detect if a row is a category/section header (mostly empty except first col)."""
    if not cells or not cells[0].strip():
        return False

    # A category row: first cell has text, most other cells are empty
    non_empty = sum(1 for c in cells[1:] if c.strip())
    return non_empty <= 1 and col_count > 2


def _smart_split_table(
    rows: list[list[str]],
    title: str,
) -> list[str]:
    """Split a large table by category rows into smaller semantic chunks."""
    if len(rows) <= _SPLIT_THRESHOLD:
        return [_rows_to_markdown_table(rows, title=title)]

    header = rows[0]
    col_count = max(len(r) for r in rows)
    chunks: list[str] = []
    current_rows: list[list[str]] = [header]
    current_title = title

    for row in rows[1:]:
        if _is_category_row(row, col_count) and len(current_rows) > 1:
            # Flush current chunk
            chunks.append(_rows_to_markdown_table(current_rows, title=current_title))
            # Start new chunk with header + category name
            category = row[0].strip()
            current_title = f"{title} — {category}"
            current_rows = [header]
        current_rows.append(row)

    # Flush last chunk
    if len(current_rows) > 1:
        chunks.append(_rows_to_markdown_table(current_rows, title=current_title))

    # If splitting didn't produce meaningful chunks, return as one table
    if len(chunks) <= 1:
        return [_rows_to_markdown_table(rows, title=title)]

    return chunks


# ------------------------------------------------------------------
# XLSX Parser
# ------------------------------------------------------------------


@register("xlsx", "xls")
class XLSXParser(BaseParser):
    """XLSX parser with formula semantics, merged cells, and smart splitting."""

    file_type = "xlsx"

    @property
    def needs_vlm(self) -> bool:
        return False

    async def to_markdown_direct(self, path: Path) -> ParsedDocument:
        from openpyxl import load_workbook

        # Need data_only=False for formulas, then a second pass for values
        wb_formulas = load_workbook(str(path), read_only=False, data_only=False)
        wb_values = load_workbook(str(path), read_only=False, data_only=True)

        pages: list[ParsedPage] = []
        page_num = 0

        for sheet_name in wb_values.sheetnames:
            ws_val = wb_values[sheet_name]
            ws_fmt = wb_formulas[sheet_name]

            # Build merged cell map
            merge_map = _build_merge_map(ws_val)

            rows: list[list[str]] = []

            for row_idx, row in enumerate(ws_val.iter_rows(min_row=1), start=1):
                cells: list[str] = []
                for col_idx, cell in enumerate(row, start=1):
                    # Check merged cell map first
                    if (row_idx, col_idx) in merge_map:
                        val = merge_map[(row_idx, col_idx)]
                    elif cell.value is not None:
                        val = str(cell.value)
                    else:
                        val = ""

                    # Add formula annotation
                    fmt_cell = ws_fmt.cell(row_idx, col_idx)
                    if (
                        fmt_cell.value
                        and isinstance(fmt_cell.value, str)
                        and fmt_cell.value.startswith("=")
                    ):
                        val += _formula_annotation(fmt_cell.value)

                    cells.append(val)

                # Skip completely empty rows
                if any(c.strip() for c in cells):
                    rows.append(cells)

            if not rows:
                continue

            # Smart split for large tables
            chunks = _smart_split_table(rows, title=sheet_name)

            for chunk_md in chunks:
                page_num += 1
                pages.append(ParsedPage(page_number=page_num, markdown=chunk_md))

            logger.debug("Sheet '%s': %d rows → %d chunk(s)", sheet_name, len(rows), len(chunks))

        wb_values.close()
        wb_formulas.close()

        return ParsedDocument(
            source_path=str(path),
            file_type="xlsx",
            pages=pages,
        )

    def extract_metadata(self, path: Path) -> dict:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True)
        meta = {
            "sheet_names": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
        }
        wb.close()
        return meta


# ---------------------------------------------------------------------------
# CSV / TSV Parser
# ---------------------------------------------------------------------------


@register("csv", "tsv")
class CSVParser(BaseParser):
    """CSV/TSV parser: direct conversion to Markdown table."""

    file_type = "csv"

    @property
    def needs_vlm(self) -> bool:
        return False

    async def to_markdown_direct(self, path: Path) -> ParsedDocument:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","

        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            rows = [row for row in reader if any(cell.strip() for cell in row)]

        md = _rows_to_markdown_table(rows, title=path.stem)

        return ParsedDocument(
            source_path=str(path),
            file_type="csv",
            pages=[ParsedPage(page_number=1, markdown=md)],
        )

    def extract_metadata(self, path: Path) -> dict:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","

        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            row_count = sum(1 for _ in reader)

        return {"row_count": row_count, "delimiter": delimiter}
