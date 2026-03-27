"""Tests for enhanced XLSX parsing: formulas, merged cells, smart split."""

import pytest

from file_parse_engine.parsers.spreadsheet import (
    _formula_annotation,
    _is_category_row,
    _smart_split_table,
    _rows_to_markdown_table,
)


class TestFormulaAnnotation:

    def test_sum(self):
        assert "sum" in _formula_annotation("=SUM(A1:A10)")

    def test_average(self):
        assert "avg" in _formula_annotation("=AVERAGE(B2:B5)")

    def test_vlookup(self):
        assert "lookup" in _formula_annotation("=VLOOKUP(A1,B:C,2,0)")

    def test_if(self):
        assert "conditional" in _formula_annotation("=IF(A1>0,1,0)")

    def test_generic(self):
        result = _formula_annotation("=A1+B1*2")
        assert "formula:" in result

    def test_no_formula(self):
        assert _formula_annotation("hello") == ""
        assert _formula_annotation("") == ""
        assert _formula_annotation("123") == ""


class TestCategoryRowDetection:

    def test_category_row(self):
        assert _is_category_row(["Operating Expenses", "", "", ""], 4)

    def test_category_with_one_value(self):
        assert _is_category_row(["Total", "100", "", ""], 4)

    def test_not_category_all_filled(self):
        assert not _is_category_row(["A", "B", "C", "D"], 4)

    def test_not_category_empty_first(self):
        assert not _is_category_row(["", "B", "C", ""], 4)

    def test_not_category_two_cols(self):
        # Two-column tables shouldn't split
        assert not _is_category_row(["Label", ""], 2)


class TestSmartSplit:

    def test_small_table_no_split(self):
        rows = [["A", "B"]] + [[str(i), str(i * 2)] for i in range(10)]
        result = _smart_split_table(rows, "Sheet1")
        assert len(result) == 1

    def test_large_table_with_categories(self):
        header = ["Item", "Q1", "Q2", "Q3"]
        rows = [header]

        # Category 1
        rows.append(["Revenue", "", "", ""])
        for i in range(20):
            rows.append([f"Product {i}", "100", "200", "300"])

        # Category 2
        rows.append(["Expenses", "", "", ""])
        for i in range(20):
            rows.append([f"Cost {i}", "50", "60", "70"])

        # Category 3
        rows.append(["Summary", "", "", ""])
        for i in range(15):
            rows.append([f"Total {i}", "150", "260", "370"])

        result = _smart_split_table(rows, "Finance")
        assert len(result) >= 2
        # Each chunk should have the header
        for chunk in result:
            assert "| Item |" in chunk

    def test_no_category_rows(self):
        header = ["A", "B", "C"]
        rows = [header] + [[str(i), str(i), str(i)] for i in range(60)]
        result = _smart_split_table(rows, "Data")
        # No splitting if no category rows detected
        assert len(result) == 1


class TestXLSXParsing:

    @pytest.mark.asyncio
    async def test_basic_xlsx(self, tmp_path):
        from openpyxl import Workbook
        from file_parse_engine.parsers.spreadsheet import XLSXParser

        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws.append(["Name", "Value"])
        ws.append(["Alice", 100])
        ws.append(["Bob", 200])
        wb.save(str(tmp_path / "test.xlsx"))

        parser = XLSXParser()
        result = await parser.to_markdown_direct(tmp_path / "test.xlsx")

        md = result.to_markdown()
        assert "Alice" in md
        assert "Bob" in md
        assert "| Name |" in md

    @pytest.mark.asyncio
    async def test_formula_annotation(self, tmp_path):
        from openpyxl import Workbook
        from file_parse_engine.parsers.spreadsheet import XLSXParser

        wb = Workbook()
        ws = wb.active
        ws.title = "Formulas"
        ws.append(["A", "B", "Total"])
        ws.append([10, 20, "=SUM(A2:B2)"])
        wb.save(str(tmp_path / "formulas.xlsx"))

        parser = XLSXParser()
        result = await parser.to_markdown_direct(tmp_path / "formulas.xlsx")

        md = result.to_markdown()
        assert "<!-- sum:" in md

    @pytest.mark.asyncio
    async def test_merged_cells(self, tmp_path):
        from openpyxl import Workbook
        from file_parse_engine.parsers.spreadsheet import XLSXParser

        wb = Workbook()
        ws = wb.active
        ws.title = "Merged"
        ws.append(["Q1", "", "Q2", ""])
        ws.merge_cells("A1:B1")
        ws.merge_cells("C1:D1")
        ws.append(["Jan", "Feb", "Mar", "Apr"])
        ws.append([10, 20, 30, 40])
        wb.save(str(tmp_path / "merged.xlsx"))

        parser = XLSXParser()
        result = await parser.to_markdown_direct(tmp_path / "merged.xlsx")

        md = result.to_markdown()
        # Q1 should appear in both merged columns
        assert md.count("Q1") >= 2
